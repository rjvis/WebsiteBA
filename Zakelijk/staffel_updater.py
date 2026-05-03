#!/usr/bin/env python3
"""
BOB Autowas — Zakelijk Staffel Updater
=======================================
Leest WebsiteBA/Zakelijk/zakelijk_staffel.xlsx
en werkt de STAFFEL in BOB_Zakelijk_Aanvraag.html automatisch bij.

Gebruik:
  python staffel_updater.py

Vereist: pip install openpyxl
"""

import sys
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installeer openpyxl eerst: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
EXCEL_PATH = SCRIPT_DIR / "Zakelijk" / "zakelijk_staffel.xlsx"
HTML_PATH  = SCRIPT_DIR / "BOB_Zakelijk_Aanvraag.html"

def main():
    if not EXCEL_PATH.exists():
        print(f"Fout: {EXCEL_PATH} niet gevonden.")
        sys.exit(1)
    if not HTML_PATH.exists():
        print(f"Fout: {HTML_PATH} niet gevonden.")
        sys.exit(1)

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Lees de staffeldata: rij 7 t/m 56 (1 t/m 50 wasbeurten)
    # Kolommen: A=n, B=Normaal, C=Intensief, D=Proteqt, E=spacer, F=IntN/I, G=IntProteqt
    staffel = {}
    for row in ws.iter_rows(min_row=7, max_row=56, values_only=True):
        n = row[0]
        if n is None or not isinstance(n, (int, float)):
            continue
        n = int(n)
        normaal    = float(row[1] or 0)
        intensief  = float(row[2] or 0)
        proteqt    = float(row[3] or 0)
        int_ni     = float(row[5] or 0)  # kolom F (index 5)
        int_prot   = float(row[6] or 0)  # kolom G (index 6)
        staffel[n] = [normaal, intensief, proteqt, int_ni, int_prot]

    if len(staffel) != 50:
        print(f"Waarschuwing: {len(staffel)} rijen gelezen, verwacht 50.")

    # Bouw de nieuwe STAFFEL JS string
    lines = ["var STAFFEL = {"]
    for n in range(1, 51):
        if n not in staffel:
            print(f"  Waarschuwing: rij {n} ontbreekt in Excel, wordt overgeslagen.")
            continue
        v = staffel[n]
        comma = "" if n == 50 else ","
        lines.append(f"  {n}: [{v[0]:.2f}, {v[1]:.2f}, {v[2]:.2f}, {v[3]:.2f}, {v[4]:.2f}]{comma}")
    lines.append("};")
    new_staffel = "\n".join(lines)

    # Vervang de STAFFEL in de HTML
    with open(HTML_PATH, encoding="utf-8") as f:
        html = f.read()

    pattern = re.compile(r"var STAFFEL = \{.*?\};", re.DOTALL)
    if not pattern.search(html):
        print("Fout: var STAFFEL niet gevonden in BOB_Zakelijk_Aanvraag.html")
        sys.exit(1)

    html = pattern.sub(new_staffel, html, count=1)

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"✓ Staffel bijgewerkt: {len(staffel)} rijen verwerkt.")
    print()
    print("Steekproef (eerste 5 rijen):")
    for n in [1, 5, 10, 20, 50]:
        v = staffel.get(n, [])
        if v:
            print(f"  n={n:>2}: Normaal={v[0]:.2f} | Intensief={v[1]:.2f} | Proteqt={v[2]:.2f} | Int.N/I={v[3]:.2f} | Int.Prot={v[4]:.2f}")

if __name__ == "__main__":
    main()
